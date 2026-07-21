from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_POST

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins

from .models import Contrato


SECRETARIAS = [
    ('administracao', 'Administração e Finanças'),
    ('assistencia', 'Assistência Social'),
    ('agricultura', 'Agricultura'),
    ('educacao', 'Educação'),
    ('infraestrutura', 'Infraestrutura'),
    ('saude', 'Saúde'),
]

MODALIDADES = [
    ('credenciamento', 'Credenciamento'),
    ('prestacao', 'Prestação de Serviço'),
]

SITUACOES = [
    ('ativo', 'Ativo'),
    ('demitido', 'Demitido'),
]


def limpar_texto(valor):
    return (valor or '').strip()


def converter_decimal(valor):
    valor = limpar_texto(valor)
    valor = (
        valor.replace('R$', '')
        .replace('.', '')
        .replace(',', '.')
        .replace(' ', '')
    )

    if not valor:
        return Decimal('0.00')

    try:
        return Decimal(valor)
    except (InvalidOperation, ValueError):
        return Decimal('0.00')


def obter_next_url_segura(request):
    next_url = request.POST.get('next') or request.GET.get('next') or ''

    if next_url and url_has_allowed_host_and_scheme(
        url=next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return next_url

    return ''


def aplicar_filtros(queryset, request):
    busca = limpar_texto(request.GET.get('busca'))
    secretaria = limpar_texto(request.GET.get('secretaria'))
    modalidade = limpar_texto(request.GET.get('modalidade'))
    situacao = limpar_texto(request.GET.get('situacao'))

    if busca:
        queryset = queryset.filter(
            Q(nome__icontains=busca)
            | Q(cpf__icontains=busca)
            | Q(rg__icontains=busca)
            | Q(funcao__icontains=busca)
            | Q(local_trabalho__icontains=busca)
            | Q(banco__icontains=busca)
        )

    if secretaria:
        queryset = queryset.filter(secretaria=secretaria)

    if modalidade:
        queryset = queryset.filter(modalidade=modalidade)

    if situacao:
        queryset = queryset.filter(situacao=situacao)

    return queryset


@login_required(login_url='/login/')
def home(request):
    ativos = Contrato.objects.filter(situacao='ativo')
    ativos_filtrados = aplicar_filtros(ativos, request)

    resumo_geral = ativos_filtrados.aggregate(
        total_funcionarios=Count('id'),
        valor_total=Coalesce(Sum('salario_fixo'), Decimal('0.00')),
    )

    resumo_credenciamento = ativos_filtrados.filter(
        modalidade='credenciamento'
    ).aggregate(
        quantidade=Count('id'),
        valor=Coalesce(Sum('salario_fixo'), Decimal('0.00')),
    )

    resumo_prestacao = ativos_filtrados.filter(
        modalidade='prestacao'
    ).aggregate(
        quantidade=Count('id'),
        valor=Coalesce(Sum('salario_fixo'), Decimal('0.00')),
    )

    resumo_secretarias = []

    for codigo, nome in SECRETARIAS:
        credenciamento = ativos_filtrados.filter(
            secretaria=codigo,
            modalidade='credenciamento',
        ).aggregate(
            quantidade=Count('id'),
            valor=Coalesce(Sum('salario_fixo'), Decimal('0.00')),
        )

        prestacao = ativos_filtrados.filter(
            secretaria=codigo,
            modalidade='prestacao',
        ).aggregate(
            quantidade=Count('id'),
            valor=Coalesce(Sum('salario_fixo'), Decimal('0.00')),
        )

        resumo_secretarias.append(
            {
                'codigo': codigo,
                'nome': nome,
                'credenciamento_quantidade': credenciamento['quantidade'],
                'credenciamento_valor': credenciamento['valor'],
                'prestacao_quantidade': prestacao['quantidade'],
                'prestacao_valor': prestacao['valor'],
                'total_quantidade': (
                    credenciamento['quantidade'] + prestacao['quantidade']
                ),
                'total_valor': credenciamento['valor'] + prestacao['valor'],
            }
        )

    contratos_recentes = (
        Contrato.objects.select_related()
        .order_by('-criado_em')[:10]
    )

    context = {
        'total_funcionarios': resumo_geral['total_funcionarios'],
        'valor_total': resumo_geral['valor_total'],
        'credenciamento_quantidade': resumo_credenciamento['quantidade'],
        'credenciamento_valor': resumo_credenciamento['valor'],
        'prestacao_quantidade': resumo_prestacao['quantidade'],
        'prestacao_valor': resumo_prestacao['valor'],
        'total_demitidos': Contrato.objects.filter(situacao='demitido').count(),
        'resumo_secretarias': resumo_secretarias,
        'contratos_recentes': contratos_recentes,
        'secretarias': SECRETARIAS,
        'modalidades': MODALIDADES,
        'secretaria_filtro': request.GET.get('secretaria', ''),
        'modalidade_filtro': request.GET.get('modalidade', ''),
        'busca': request.GET.get('busca', ''),
        'grafico_secretarias': [item['nome'] for item in resumo_secretarias],
        'grafico_credenciamento_quantidades': [
            item['credenciamento_quantidade'] for item in resumo_secretarias
        ],
        'grafico_prestacao_quantidades': [
            item['prestacao_quantidade'] for item in resumo_secretarias
        ],
        'grafico_credenciamento_valores': [
            float(item['credenciamento_valor']) for item in resumo_secretarias
        ],
        'grafico_prestacao_valores': [
            float(item['prestacao_valor']) for item in resumo_secretarias
        ],
    }

    return render(request, 'contratos/dashboard.html', context)


@login_required(login_url='/login/')
def cadastrar_contrato(request):
    if request.method == 'POST':
        cpf = limpar_texto(request.POST.get('cpf'))

        if Contrato.objects.filter(cpf=cpf).exists():
            messages.error(request, 'Já existe um servidor cadastrado com este CPF.')
            return render(
                request,
                'contratos/formulario.html',
                {
                    'secretarias': SECRETARIAS,
                    'modalidades': MODALIDADES,
                    'situacoes': SITUACOES,
                    'dados': request.POST,
                    'titulo': 'Cadastrar servidor',
                },
            )

        contrato = Contrato.objects.create(
            nome=limpar_texto(request.POST.get('nome')),
            funcao=limpar_texto(request.POST.get('funcao')),
            local_trabalho=limpar_texto(request.POST.get('local_trabalho')),
            cpf=cpf,
            rg=limpar_texto(request.POST.get('rg')),
            data_nascimento=request.POST.get('data_nascimento') or None,
            banco=limpar_texto(request.POST.get('banco')),
            agencia=limpar_texto(request.POST.get('agencia')),
            conta_bancaria=limpar_texto(request.POST.get('conta_bancaria')),
            salario_fixo=converter_decimal(request.POST.get('salario_fixo')),
            secretaria=limpar_texto(request.POST.get('secretaria')),
            modalidade=limpar_texto(request.POST.get('modalidade')),
            situacao=limpar_texto(request.POST.get('situacao')) or 'ativo',
        )

        messages.success(request, f'{contrato.nome} foi cadastrado com sucesso.')
        return redirect('listar_contratos')

    return render(
        request,
        'contratos/formulario.html',
        {
            'secretarias': SECRETARIAS,
            'modalidades': MODALIDADES,
            'situacoes': SITUACOES,
            'titulo': 'Cadastrar servidor',
        },
    )


@login_required(login_url='/login/')
def listar_contratos(request):
    contratos = aplicar_filtros(Contrato.objects.all(), request)

    ordem = request.GET.get('ordem', 'nome')
    ordenacoes = {
        'nome': 'nome',
        '-nome': '-nome',
        'recente': '-criado_em',
        'antigo': 'criado_em',
        'maior_salario': '-salario_fixo',
        'menor_salario': 'salario_fixo',
    }
    contratos = contratos.order_by(ordenacoes.get(ordem, 'nome'))

    totais = contratos.aggregate(
        quantidade=Count('id'),
        valor=Coalesce(Sum('salario_fixo'), Decimal('0.00')),
    )

    context = {
        'contratos': contratos,
        'total_resultados': totais['quantidade'],
        'valor_resultados': totais['valor'],
        'secretarias': SECRETARIAS,
        'modalidades': MODALIDADES,
        'situacoes': SITUACOES,
        'busca': request.GET.get('busca', ''),
        'secretaria_filtro': request.GET.get('secretaria', ''),
        'modalidade_filtro': request.GET.get('modalidade', ''),
        'situacao_filtro': request.GET.get('situacao', ''),
        'ordem': ordem,
    }

    return render(request, 'contratos/lista.html', context)


@login_required(login_url='/login/')
def contratos_por_categoria(request, modalidade, secretaria):
    modalidades_validas = {codigo for codigo, _ in MODALIDADES}
    secretarias_validas = {codigo for codigo, _ in SECRETARIAS}

    if modalidade not in modalidades_validas or secretaria not in secretarias_validas:
        messages.error(request, 'Categoria de contratos inválida.')
        return redirect('listar_contratos')

    contratos = Contrato.objects.filter(
        modalidade=modalidade,
        secretaria=secretaria,
    )
    contratos = aplicar_filtros(contratos, request).order_by('nome')

    modalidade_nome = dict(MODALIDADES)[modalidade]
    secretaria_nome = dict(SECRETARIAS)[secretaria]

    totais = contratos.aggregate(
        quantidade=Count('id'),
        valor=Coalesce(Sum('salario_fixo'), Decimal('0.00')),
    )

    return render(
        request,
        'contratos/lista.html',
        {
            'contratos': contratos,
            'total_resultados': totais['quantidade'],
            'valor_resultados': totais['valor'],
            'secretarias': SECRETARIAS,
            'modalidades': MODALIDADES,
            'situacoes': SITUACOES,
            'modalidade_filtro': modalidade,
            'secretaria_filtro': secretaria,
            'situacao_filtro': request.GET.get('situacao', ''),
            'busca': request.GET.get('busca', ''),
            'titulo_lista': f'{modalidade_nome} — {secretaria_nome}',
        },
    )


@login_required(login_url='/login/')
def detalhe_contrato(request, id):
    contrato = get_object_or_404(Contrato, id=id)
    return render(request, 'contratos/detalhe.html', {'contrato': contrato})


@login_required(login_url='/login/')
def editar_contrato(request, id):
    contrato = get_object_or_404(Contrato, id=id)
    next_url = obter_next_url_segura(request)

    if request.method == 'POST':
        cpf = limpar_texto(request.POST.get('cpf'))

        cpf_duplicado = (
            Contrato.objects.filter(cpf=cpf)
            .exclude(id=contrato.id)
            .exists()
        )

        if cpf_duplicado:
            messages.error(request, 'Já existe outro servidor com este CPF.')
            return render(
                request,
                'contratos/formulario.html',
                {
                    'contrato': contrato,
                    'secretarias': SECRETARIAS,
                    'modalidades': MODALIDADES,
                    'situacoes': SITUACOES,
                    'dados': request.POST,
                    'titulo': 'Editar servidor',
                    'next_url': next_url,
                },
            )

        contrato.nome = limpar_texto(request.POST.get('nome'))
        contrato.funcao = limpar_texto(request.POST.get('funcao'))
        contrato.local_trabalho = limpar_texto(request.POST.get('local_trabalho'))
        contrato.cpf = cpf
        contrato.rg = limpar_texto(request.POST.get('rg'))
        contrato.data_nascimento = request.POST.get('data_nascimento') or None
        contrato.banco = limpar_texto(request.POST.get('banco'))
        contrato.agencia = limpar_texto(request.POST.get('agencia'))
        contrato.conta_bancaria = limpar_texto(request.POST.get('conta_bancaria'))
        contrato.salario_fixo = converter_decimal(request.POST.get('salario_fixo'))
        contrato.secretaria = limpar_texto(request.POST.get('secretaria'))
        contrato.modalidade = limpar_texto(request.POST.get('modalidade'))
        contrato.situacao = limpar_texto(request.POST.get('situacao')) or 'ativo'
        contrato.save()

        messages.success(request, 'Cadastro atualizado com sucesso.')
        return redirect(next_url or 'listar_contratos')

    return render(
        request,
        'contratos/formulario.html',
        {
            'contrato': contrato,
            'secretarias': SECRETARIAS,
            'modalidades': MODALIDADES,
            'situacoes': SITUACOES,
            'titulo': 'Editar servidor',
            'next_url': next_url,
        },
    )


@login_required(login_url='/login/')
@require_POST
def alterar_situacao(request, id):
    contrato = get_object_or_404(Contrato, id=id)
    situacao = request.POST.get('situacao')

    situacoes_validas = {codigo for codigo, _ in SITUACOES}

    if situacao in situacoes_validas:
        contrato.situacao = situacao
        contrato.save(update_fields=['situacao', 'atualizado_em'])
        messages.success(request, 'Situação atualizada com sucesso.')
    else:
        messages.error(request, 'Situação inválida.')

    return redirect(obter_next_url_segura(request) or 'listar_contratos')


@login_required(login_url='/login/')
@require_POST
def excluir_contrato(request, id):
    contrato = get_object_or_404(Contrato, id=id)
    nome = contrato.nome
    contrato.delete()
    messages.success(request, f'O cadastro de {nome} foi excluído.')
    return redirect(obter_next_url_segura(request) or 'listar_contratos')


@login_required(login_url='/login/')
def relatorios(request):
    contratos = aplicar_filtros(Contrato.objects.all(), request)

    resumo_secretaria = (
        contratos.values('secretaria')
        .annotate(
            quantidade=Count('id'),
            valor_total=Coalesce(Sum('salario_fixo'), Decimal('0.00')),
        )
        .order_by('secretaria')
    )

    resumo_modalidade = (
        contratos.values('modalidade')
        .annotate(
            quantidade=Count('id'),
            valor_total=Coalesce(Sum('salario_fixo'), Decimal('0.00')),
        )
        .order_by('modalidade')
    )

    return render(
        request,
        'contratos/relatorios.html',
        {
            'resumo_secretaria': resumo_secretaria,
            'resumo_modalidade': resumo_modalidade,
            'secretarias': SECRETARIAS,
            'modalidades': MODALIDADES,
            'situacoes': SITUACOES,
            'busca': request.GET.get('busca', ''),
            'secretaria_filtro': request.GET.get('secretaria', ''),
            'modalidade_filtro': request.GET.get('modalidade', ''),
            'situacao_filtro': request.GET.get('situacao', ''),
        },
    )


@login_required(login_url='/login/')
def exportar_contratos_excel(request):
    contratos = aplicar_filtros(Contrato.objects.all(), request).order_by(
        'secretaria',
        'modalidade',
        'nome',
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Contratos'

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.freeze_panes = 'A4'

    ws.page_margins = PageMargins(
        left=0.15,
        right=0.15,
        top=0.20,
        bottom=0.20,
        header=0,
        footer=0,
    )

    ws.merge_cells('A1:N1')
    titulo = ws['A1']
    titulo.value = 'RELAÇÃO DE CONTRATOS — PREFEITURA MUNICIPAL'
    titulo.font = Font(bold=True, color='FFFFFF', size=14)
    titulo.fill = PatternFill('solid', fgColor='17365D')
    titulo.alignment = Alignment(horizontal='center', vertical='center')

    filtros = []
    if request.GET.get('secretaria'):
        filtros.append(
            f"Secretaria: {dict(SECRETARIAS).get(request.GET['secretaria'], request.GET['secretaria'])}"
        )
    if request.GET.get('modalidade'):
        filtros.append(
            f"Modalidade: {dict(MODALIDADES).get(request.GET['modalidade'], request.GET['modalidade'])}"
        )
    if request.GET.get('situacao'):
        filtros.append(
            f"Situação: {dict(SITUACOES).get(request.GET['situacao'], request.GET['situacao'])}"
        )
    if request.GET.get('busca'):
        filtros.append(f"Pesquisa: {request.GET['busca']}")

    ws.merge_cells('A2:N2')
    ws['A2'] = ' | '.join(filtros) if filtros else 'Todos os contratos'
    ws['A2'].font = Font(italic=True, color='475569', size=10)
    ws['A2'].alignment = Alignment(horizontal='center')

    cabecalhos = [
        'NOME',
        'FUNÇÃO',
        'LOCAL DE TRABALHO',
        'CPF',
        'RG',
        'DATA DE NASCIMENTO',
        'BANCO',
        'AGÊNCIA',
        'CONTA',
        'SALÁRIO FIXO',
        'SECRETARIA',
        'MODALIDADE',
        'SITUAÇÃO',
        'CADASTRADO EM',
    ]
    ws.append(cabecalhos)

    borda = Border(
        left=Side(style='thin', color='94A3B8'),
        right=Side(style='thin', color='94A3B8'),
        top=Side(style='thin', color='94A3B8'),
        bottom=Side(style='thin', color='94A3B8'),
    )

    for cell in ws[3]:
        cell.font = Font(bold=True, color='FFFFFF', size=9)
        cell.fill = PatternFill('solid', fgColor='F97316')
        cell.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True,
        )
        cell.border = borda

    for contrato in contratos:
        ws.append(
            [
                contrato.nome,
                contrato.funcao,
                contrato.local_trabalho,
                contrato.cpf,
                contrato.rg,
                contrato.data_nascimento.strftime('%d/%m/%Y')
                if contrato.data_nascimento
                else '',
                contrato.banco,
                contrato.agencia,
                contrato.conta_bancaria,
                contrato.salario_fixo,
                contrato.get_secretaria_display(),
                contrato.get_modalidade_display(),
                contrato.get_situacao_display(),
                timezone.localtime(contrato.criado_em).strftime('%d/%m/%Y %H:%M'),
            ]
        )

        linha = ws.max_row
        for cell in ws[linha]:
            cell.font = Font(size=9)
            cell.border = borda
            cell.alignment = Alignment(vertical='center', wrap_text=True)

        ws.cell(row=linha, column=10).number_format = 'R$ #,##0.00'

    larguras = {
        'A': 30,
        'B': 24,
        'C': 28,
        'D': 16,
        'E': 15,
        'F': 18,
        'G': 18,
        'H': 12,
        'I': 18,
        'J': 16,
        'K': 28,
        'L': 24,
        'M': 14,
        'N': 20,
    }

    for coluna, largura in larguras.items():
        ws.column_dimensions[coluna].width = largura

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 30

    if ws.max_row >= 4:
        ws.auto_filter.ref = f'A3:N{ws.max_row}'

    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    )

    data_atual = timezone.localdate().strftime('%d-%m-%Y')
    response['Content-Disposition'] = (
        f'attachment; filename="contratos_{data_atual}.xlsx"'
    )

    wb.save(response)
    return response